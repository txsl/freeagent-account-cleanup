import json
import shutil
import subprocess
import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

import build_review_workbook as brw


class TestEnrich:
    def test_no_explanations_returns_blank_flags(self):
        txn = {"bank_transaction_explanations": []}
        result = brw.enrich(txn)
        assert result == {
            "explained": "N", "explanation_type": "",
            "is_deletable": "", "is_locked": "", "locked_reason": "",
        }

    def test_missing_explanations_key_treated_as_none(self):
        assert brw.enrich({})["explained"] == "N"

    def test_explained_deletable_uses_type_field(self):
        txn = {"bank_transaction_explanations": [{"url": "https://api.freeagent.com/v2/bank_transaction_explanations/1"}]}
        with patch("build_review_workbook.fa.get_explanation") as mock_get:
            mock_get.return_value = {"type": "Payment", "is_deletable": True, "is_locked": False, "locked_reason": ""}
            result = brw.enrich(txn)
        assert result["explained"] == "Y"
        assert result["explanation_type"] == "Payment"
        assert result["is_deletable"] == "Y"
        assert result["is_locked"] == "N"

    def test_falls_back_to_entry_type_when_type_key_absent(self):
        # FreeAgent's docs list the attribute as `type` but at least one
        # documented example embeds it as `entry_type` — enrich() must
        # handle both.
        txn = {"bank_transaction_explanations": [{"url": "..."}]}
        with patch("build_review_workbook.fa.get_explanation") as mock_get:
            mock_get.return_value = {"entry_type": "Invoice Receipt", "is_deletable": True}
            result = brw.enrich(txn)
        assert result["explanation_type"] == "Invoice Receipt"

    def test_not_deletable_and_locked_surface_reason(self):
        txn = {"bank_transaction_explanations": [{"url": "..."}]}
        with patch("build_review_workbook.fa.get_explanation") as mock_get:
            mock_get.return_value = {
                "type": "Payment", "is_deletable": False,
                "is_locked": True, "locked_reason": "VAT return filed",
            }
            result = brw.enrich(txn)
        assert result["is_deletable"] == "N"
        assert result["is_locked"] == "Y"
        assert result["locked_reason"] == "VAT return filed"

    def test_multiple_explanations_flags_count_for_manual_check(self):
        txn = {"bank_transaction_explanations": [{"url": "..1"}, {"url": "..2"}]}
        with patch("build_review_workbook.fa.get_explanation") as mock_get:
            mock_get.return_value = {"type": "Payment", "is_deletable": True}
            result = brw.enrich(txn)
        assert "more explanation" in result["explanation_type"]


class TestFetchAndCache:
    def test_builds_rows_and_writes_json_cache(self, tmp_path, monkeypatch):
        monkeypatch.setattr(brw, "CACHE_DIR", str(tmp_path / "cache"))
        fake_txns = [{
            "url": "https://api.freeagent.com/v2/bank_transactions/1",
            "amount": "-42.50",
            "dated_on": "2026-02-01",
            "description": "Test txn",
            "full_description": "Test txn full",
            "uploaded_at": "2026-02-01T00:00:00.000Z",
            "is_manual": False,
            "bank_transaction_explanations": [],
        }]
        with patch("build_review_workbook.fa.get_bank_transactions", return_value=fake_txns):
            rows = brw.fetch_and_cache("99", "polluted")

        assert len(rows) == 1
        assert rows[0]["amount"] == -42.5  # string -> float conversion
        assert rows[0]["explained"] == "N"

        cache_file = tmp_path / "cache" / "polluted_raw.json"
        assert cache_file.exists()
        assert json.loads(cache_file.read_text()) == fake_txns


def _row(url, dated_on, amount, description):
    return dict(
        url=url, dated_on=dated_on, amount=amount, description=description,
        full_description="", is_manual=False, uploaded_at="",
        explained="N", explanation_type="", is_deletable="", is_locked="", locked_reason="",
    )


def _build_test_workbook(path):
    """Shared fixture data for the formula tests below: one unmatched
    transaction, one clean 1:1 match, and a same-day/same-amount pair that
    should come out ambiguous on both sides."""
    polluted_rows = [
        _row("u1", "2026-01-05", -50.0, "Coffee supplier"),   # no counterpart -> unmatched
        _row("u2", "2026-01-06", -200.0, "Duplicate rent"),   # exactly one counterpart -> matched
        _row("u3", "2026-01-08", -100.0, "Ambiguous A"),      # two same-day/same-amount -> ambiguous
        _row("u4", "2026-01-08", -100.0, "Ambiguous B"),
    ]
    clean_rows = [
        _row("c1", "2026-01-06", -200.0, "Duplicate rent"),
        _row("c2", "2026-01-08", -100.0, "Ambiguous A"),
        _row("c3", "2026-01-08", -100.0, "Ambiguous C"),
    ]

    wb = openpyxl.Workbook()
    ws_polluted = wb.active
    ws_polluted.title = "Account A (Polluted)"
    ws_clean = wb.create_sheet("Account B (Clean)")
    brw.write_raw_sheet(ws_polluted, polluted_rows)
    brw.write_raw_sheet(ws_clean, clean_rows)
    brw.add_match_and_approval_columns(ws_polluted, len(polluted_rows), "Account B (Clean)")
    wb.save(path)
    return ws_polluted


class TestSheetStructureWithoutAnEngine:
    """Pure openpyxl checks — no spreadsheet engine required, so these run
    on every machine regardless of whether LibreOffice is installed."""

    def test_approve_delete_column_exists_and_starts_blank(self, tmp_path):
        path = tmp_path / "test_review.xlsx"
        ws = _build_test_workbook(path)
        headers = [c.value for c in ws[1]]
        assert "approve_delete" in headers
        approve_col = headers.index("approve_delete") + 1
        assert ws.cell(row=2, column=approve_col).value is None

    def test_match_formula_text_is_exactly_as_expected(self, tmp_path):
        # A weaker check than actually recalculating (see the class below),
        # but it needs no external engine, so it's a useful floor: if
        # someone edits the formula-building code and gets the cell
        # references or function names wrong, this catches it immediately
        # even on a machine without LibreOffice.
        path = tmp_path / "test_review.xlsx"
        ws = _build_test_workbook(path)
        headers = [c.value for c in ws[1]]
        count_col = headers.index("match_count") + 1
        status_col = headers.index("match_status") + 1

        assert ws.cell(row=2, column=count_col).value == (
            "=COUNTIFS('Account B (Clean)'!$B:$B,B2,'Account B (Clean)'!$C:$C,C2)"
        )
        assert ws.cell(row=2, column=status_col).value == (
            '=IF(M2=0,"unmatched",IF(M2=1,"matched","AMBIGUOUS - review"))'
        )


class TestBalanceReconciliation:
    def test_formula_text_and_layout_without_an_engine(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instructions"
        next_row = brw.add_balance_reconciliation(ws, 1, 950.0, "Account A (Polluted)", "Account B (Clean)")

        # balance value, sum formula, and projected-balance formula land on
        # the three rows we expect, each with its label in column A.
        assert ws.cell(row=1, column=1).value == "Balance reconciliation"
        assert ws.cell(row=2, column=2).value == 950.0
        assert ws.cell(row=3, column=3).value == (
            "=SUMIFS('Account A (Polluted)'!$C:$C,"
            "'Account A (Polluted)'!$N:$N,\"matched\")"
        )
        assert ws.cell(row=4, column=3).value == "=B2-C3"
        assert ws.cell(row=5, column=2).value == (
            "=SUMIFS('Account A (Polluted)'!$C:$C,"
            "'Account A (Polluted)'!$O:$O,\"Y\","
            "'Account A (Polluted)'!$N:$N,\"matched\")"
        )
        assert ws.cell(row=6, column=2).value == "=B2-B5"
        assert ws.cell(row=8, column=2).value == "=SUM('Account B (Clean)'!$C:$C)"
        # returns the next free row so callers can't accidentally overlap content
        assert next_row == 9

    def test_amount_and_flag_columns_referenced_match_actual_sheet_layout(self):
        # Regression guard: the C/N/O column references above are computed
        # from RAW_HEADERS + the same offsets add_match_and_approval_columns
        # uses, not hardcoded — this pins down that they still agree.
        assert get_column_letter_for_header("amount") == "C"
        assert get_column_letter_for_header_offset(2) == "N"  # match_status
        assert get_column_letter_for_header_offset(3) == "O"  # approve_delete


def get_column_letter_for_header(name):
    from openpyxl.utils import get_column_letter
    return get_column_letter(brw.RAW_HEADERS.index(name) + 1)


def get_column_letter_for_header_offset(offset):
    from openpyxl.utils import get_column_letter
    return get_column_letter(len(brw.RAW_HEADERS) + offset)


def _soffice_binary():
    return shutil.which("soffice") or shutil.which("libreoffice")


_RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def recalculate_with_libreoffice(xlsx_path, timeout=60):
    """Forces LibreOffice to recalculate every formula in xlsx_path and
    save it back in place. Runs against a throwaway user profile (created
    fresh in a temp dir) so it can't collide with, or get confused by, a
    real LibreOffice profile on the machine running the tests.

    Deliberately self-contained (no dependency on Anthropic's internal
    xlsx-skill tooling) so it works the same way in CI or on a
    contributor's laptop as it does here.
    """
    soffice = _soffice_binary()
    if not soffice:
        raise RuntimeError("soffice/libreoffice not found on PATH")

    abs_path = str(Path(xlsx_path).resolve())

    with tempfile.TemporaryDirectory(prefix="recalc-lo-profile-") as profile_dir:
        profile_url = Path(profile_dir).as_uri()

        # One-off invocation to materialize a fresh profile directory.
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init", f"-env:UserInstallation={profile_url}"],
            capture_output=True, timeout=timeout, check=True,
        )

        macro_dir = Path(profile_dir) / "user" / "basic" / "Standard"
        macro_dir.mkdir(parents=True, exist_ok=True)
        (macro_dir / "Module1.xba").write_text(_RECALCULATE_MACRO)

        subprocess.run(
            [
                soffice, "--headless", "--norestore",
                f"-env:UserInstallation={profile_url}",
                "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
                abs_path,
            ],
            capture_output=True, timeout=timeout, check=True,
        )


@pytest.mark.skipif(
    _soffice_binary() is None,
    reason="LibreOffice (soffice) not found on PATH — install it to run this check",
)
class TestMatchFormulasAgainstRealExcelEngine:
    """Writes an actual workbook and recalculates it with LibreOffice, the
    same way build_review_workbook's output would be opened by the user —
    so a broken formula shows up as a real test failure, not just a string
    comparison against formula text (see test_match_formula_text_is_exactly_as_expected
    above for the engine-free fallback version of this check)."""

    def test_unmatched_one_to_one_and_ambiguous_cases(self, tmp_path):
        path = tmp_path / "test_review.xlsx"
        _build_test_workbook(path)

        recalculate_with_libreoffice(path)

        wb2 = openpyxl.load_workbook(path, data_only=True)
        ws2 = wb2["Account A (Polluted)"]
        headers = [c.value for c in ws2[1]]
        idx = {h: i for i, h in enumerate(headers)}
        by_desc = {
            row[idx["description"]]: (row[idx["match_count"]], row[idx["match_status"]])
            for row in ws2.iter_rows(min_row=2, values_only=True)
        }

        assert by_desc["Coffee supplier"] == (0, "unmatched")
        assert by_desc["Duplicate rent"] == (1, "matched")
        assert by_desc["Ambiguous A"] == (2, "AMBIGUOUS - review")
        assert by_desc["Ambiguous B"] == (2, "AMBIGUOUS - review")

    def test_projected_balance_matches_hand_calculated_value(self, tmp_path):
        # Approved for deletion: -200 (rent) and -30 (coffee) = -230 signed
        # total (FreeAgent amounts are signed: negative = money out).
        # "Duplicate software" (-75) is matched but deliberately left
        # un-approved, to prove the formula only sums approved rows.
        #
        # Deleting fictitious *outflows* should RAISE the balance back
        # toward the true figure, so: projected = balance - sum
        #                                        = 1000.00 - (-230.00)
        #                                        = 1230.00
        # (subtracting a negative sum of erroneous debits adds it back).
        polluted_rows = [
            _row("u1", "2026-01-05", -50.0, "Coffee supplier"),   # unmatched, untouched
            _row("u2", "2026-01-06", -200.0, "Duplicate rent"),   # matched + approved
            _row("u3", "2026-01-07", -75.0, "Duplicate software"),  # matched, NOT approved
            _row("u4", "2026-01-08", -30.0, "Duplicate coffee"),  # matched + approved
        ]
        clean_rows = [
            _row("c1", "2026-01-06", -200.0, "Duplicate rent"),
            _row("c2", "2026-01-07", -75.0, "Duplicate software"),
            _row("c3", "2026-01-08", -30.0, "Duplicate coffee"),
        ]

        wb = openpyxl.Workbook()
        ws_polluted = wb.active
        ws_polluted.title = "Account A (Polluted)"
        ws_clean = wb.create_sheet("Account B (Clean)")
        brw.write_raw_sheet(ws_polluted, polluted_rows)
        brw.write_raw_sheet(ws_clean, clean_rows)
        brw.add_match_and_approval_columns(ws_polluted, len(polluted_rows), "Account B (Clean)")

        headers = [c.value for c in ws_polluted[1]]
        approve_col = headers.index("approve_delete") + 1
        desc_col = headers.index("description") + 1
        for r in range(2, ws_polluted.max_row + 1):
            desc = ws_polluted.cell(row=r, column=desc_col).value
            if desc in ("Duplicate rent", "Duplicate coffee"):  # NOT "Duplicate software"
                ws_polluted.cell(row=r, column=approve_col, value="Y")

        ws_instructions = wb.create_sheet("Instructions", 0)
        brw.add_balance_reconciliation(ws_instructions, 1, 1000.00, "Account A (Polluted)", "Account B (Clean)")

        path = tmp_path / "test_review_balance.xlsx"
        wb.save(path)
        recalculate_with_libreoffice(path)

        wb2 = openpyxl.load_workbook(path, data_only=True)
        ws2 = wb2["Instructions"]
        assert ws2.cell(row=3, column=2).value == pytest.approx(-230.0)  # signed sum: -200 + -30
        assert ws2.cell(row=4, column=2).value == pytest.approx(1230.0)  # 1000.00 - (-230.00)
