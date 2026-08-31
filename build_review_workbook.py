"""
build_review_workbook.py — pulls both bank accounts and builds an Excel
workbook for manual review before any deletion happens.

Usage:
    python build_review_workbook.py

Reads ACCOUNT_POLLUTED_ID / ACCOUNT_CLEAN_ID from .env. Writes:
    cache/polluted_raw.json, cache/clean_raw.json   — frozen "before" snapshot
    review_<timestamp>.xlsx                          — what you actually review

Sheet "Account A (Polluted)" gets two extra formula columns:
    match_count  = COUNTIFS against Account B's dated_on + amount
    match_status = "unmatched" / "matched" / "AMBIGUOUS - review"
                  (matched only when the date/amount is unique on both sides)
Plain COUNTIFS/IF only — no array formulas, FILTER, or TEXTJOIN — so every
formula is a single well-known function you can click on and read.

Workflow: filter Sheet A to match_status = "matched", sanity-check a few,
mark approve_delete = Y on the rows you're confident about, save, then run
delete_transactions.py against this file.

The Instructions tab also shows a live balance reconciliation: FreeAgent's
current balance for Account A minus the sum of whatever you've approved for
deletion so far, so you can check that number against your real bank
statement before actually deleting anything. It includes a blank target-balance
input and an alignment check for the user to fill in.
"""
import json
import os
import sys
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import freeagent_client as fa

CACHE_DIR = "cache"

# Column order matters: dated_on must be column B and amount column C on
# both sheets, since the match formulas hardcode those references.
RAW_HEADERS = [
    "url", "dated_on", "amount", "description", "full_description",
    "is_manual", "uploaded_at", "explained", "explanation_type",
    "is_deletable", "is_locked", "locked_reason",
]


def enrich(txn):
    """Pull live explanation detail for a transaction, if it has one."""
    explanations = txn.get("bank_transaction_explanations") or []
    if not explanations:
        return {
            "explained": "N", "explanation_type": "",
            "is_deletable": "", "is_locked": "", "locked_reason": "",
        }
    detail = fa.get_explanation(explanations[0]["url"])
    # FreeAgent's docs list this attribute as `type` but at least one
    # documented example embeds it as `entry_type` — check both rather
    # than trust one.
    exp_type = detail.get("type") or detail.get("entry_type") or ""
    if len(explanations) > 1:
        exp_type += f" (+{len(explanations) - 1} more explanation(s) — check manually)"
    return {
        "explained": "Y",
        "explanation_type": exp_type,
        "is_deletable": "Y" if detail.get("is_deletable") else "N",
        "is_locked": "Y" if detail.get("is_locked") else "N",
        "locked_reason": detail.get("locked_reason") or "",
    }


def fetch_and_cache(account_id, label, from_cache=False):
    rows_cache_path = f"{CACHE_DIR}/{label}_rows.json"

    if from_cache:
        if not os.path.exists(rows_cache_path):
            sys.exit(f"No cache at {rows_cache_path} — run once without --from-cache first.")
        print(f"Loading {label} account from cache ({rows_cache_path})...")
        with open(rows_cache_path) as f:
            return json.load(f)

    account_url = f"{fa.API_BASE}/bank_accounts/{account_id}"
    print(f"Fetching {label} account ({account_id})...")
    raw = fa.get_bank_transactions(account_url)

    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(f"{CACHE_DIR}/{label}_raw.json", "w") as f:
        json.dump(raw, f, indent=2)

    print(f"  {len(raw)} transactions. Pulling explanation detail...")
    rows = []
    for i, txn in enumerate(raw):
        info = enrich(txn)
        rows.append({
            "url": txn["url"],
            "dated_on": txn["dated_on"],
            "amount": float(txn["amount"]),
            "description": txn.get("description", ""),
            "full_description": txn.get("full_description", ""),
            "is_manual": txn.get("is_manual", False),
            "uploaded_at": txn.get("uploaded_at", ""),
            **info,
        })
        if (i + 1) % 25 == 0:
            print(f"    {i + 1}/{len(raw)}")

    with open(rows_cache_path, "w") as f:
        json.dump(rows, f, indent=2)
    return rows


def get_current_balance(account_id, from_cache=False):
    balance_cache_path = f"{CACHE_DIR}/{account_id}_balance.json"
    if from_cache:
        with open(balance_cache_path) as f:
            return json.load(f)["current_balance"]
    os.makedirs(CACHE_DIR, exist_ok=True)
    balance = float(fa.get_bank_account(account_id)["current_balance"])
    with open(balance_cache_path, "w") as f:
        json.dump({"current_balance": balance}, f)
    return balance


def write_raw_sheet(ws, rows):
    bold = Font(bold=True)
    for col, header in enumerate(RAW_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header).font = bold
    for r, row in enumerate(rows, start=2):
        for c, header in enumerate(RAW_HEADERS, start=1):
            ws.cell(row=r, column=c, value=row.get(header, ""))
    for c in range(1, len(RAW_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35


def add_match_and_approval_columns(ws, n_rows, other_sheet_name):
    bold = Font(bold=True)
    col_count = len(RAW_HEADERS) + 1
    col_status = len(RAW_HEADERS) + 2
    col_approve = len(RAW_HEADERS) + 3

    ws.cell(row=1, column=col_count, value="match_count").font = bold
    ws.cell(row=1, column=col_status, value="match_status").font = bold
    ws.cell(row=1, column=col_approve, value="approve_delete").font = bold
    ws.column_dimensions[get_column_letter(col_count)].width = 12
    ws.column_dimensions[get_column_letter(col_status)].width = 26
    ws.column_dimensions[get_column_letter(col_approve)].width = 16

    other = f"'{other_sheet_name}'"
    count_letter = get_column_letter(col_count)
    for r in range(2, n_rows + 2):
        ws.cell(
            row=r, column=col_count,
            value=f"=COUNTIFS({other}!$B:$B,B{r},{other}!$C:$C,C{r})",
        )
        ws.cell(
            row=r, column=col_status,
            value=(
                f'=IF({count_letter}{r}=0,"unmatched",'
                f'IF(AND({count_letter}{r}=1,'
                f'COUNTIFS($B:$B,B{r},$C:$C,C{r})=1),'
                f'"matched","AMBIGUOUS - review"))'
            ),
        )
        # approve_delete left blank for you to fill in


def add_balance_reconciliation(ws, start_row, polluted_current_balance, polluted_sheet_name, clean_sheet_name):
    """Writes a small reconciliation summary: FreeAgent's reported current
    balance for the polluted account, the sum of whatever you've approved
    for deletion so far, and the projected balance after deletion — so you
    can check that number against your real bank statement. Column
    references are computed from RAW_HEADERS rather than hardcoded, so
    they can't silently drift out of sync if the sheet layout changes.

    Returns the row number after the last row written.
    """
    amount_col = get_column_letter(RAW_HEADERS.index("amount") + 1)
    status_col = get_column_letter(len(RAW_HEADERS) + 2)   # matches add_match_and_approval_columns
    approve_col = get_column_letter(len(RAW_HEADERS) + 3)  # matches add_match_and_approval_columns
    currency_format = '_(* #,##0.00_);_(* \\ (#,##0.00);_(* "-"??_);_(@_)'
    count_format = '_(* #,##0_);_(* \\ (#,##0);_(* "-"??_);_(@_)'

    r = start_row
    ws.cell(row=r, column=1, value="Balance reconciliation").font = Font(bold=True, size=13)
    r += 1

    ws.cell(row=r, column=1, value="FreeAgent's current balance for Account A (Polluted), as of this fetch:")
    ws.cell(row=r, column=2, value=polluted_current_balance)
    ws.cell(row=r, column=2).number_format = currency_format
    balance_row = r
    r += 1

    ws.cell(row=r, column=1, value="Sum of amounts potentially duplicated (match_status = matched):")
    ws.cell(row=r, column=3, value=(
        f"=SUMIFS('{polluted_sheet_name}'!${amount_col}:${amount_col},"
        f"'{polluted_sheet_name}'!${status_col}:${status_col},\"matched\")"
    ))
    ws.cell(row=r, column=3).number_format = currency_format
    sum_row = r
    r += 1

    label_cell = ws.cell(row=r, column=1, value="Projected balance after deleting matched rows:")
    value_cell = ws.cell(row=r, column=3, value=f"=B{balance_row}-C{sum_row}")
    value_cell.number_format = currency_format
    r += 1

    ws.cell(row=r, column=1, value="Sum of amounts approved for deletion (approve_delete = Y and match_status = matched):")
    ws.cell(row=r, column=2, value=(
        f"=SUMIFS('{polluted_sheet_name}'!${amount_col}:${amount_col},"
        f"'{polluted_sheet_name}'!${approve_col}:${approve_col},\"Y\","
        f"'{polluted_sheet_name}'!${status_col}:${status_col},\"matched\")"
    ))
    ws.cell(row=r, column=2).number_format = currency_format
    sum_row = r
    r += 1

    label_cell = ws.cell(row=r, column=1, value="Projected balance after deleting the approved rows:")
    value_cell = ws.cell(row=r, column=2, value=f"=B{balance_row}-B{sum_row}")
    approved_projected_balance_row = r
    value_cell.number_format = currency_format
    label_cell.font = Font(bold=True)
    value_cell.font = Font(bold=True)
    r += 1

    # Leave a blank spacer row before the explanatory note.
    r += 1
    ws.cell(
        row=r, column=1,
        value=(
            "Compare the projected balance above to your real bank statement balance as of "
            "today. If they match, that's strong evidence the approved rows are exactly the "
            "duplicates and nothing else. If they don't, something's off — re-check before "
            "deleting anything."
        ),
    )
    r += 1

    # Leave a blank spacer row before the user-entered target section.
    target_balance_row = r
    target_label = ws.cell(row=r, column=1, value="Target balance (User fills in)")
    target_value = ws.cell(row=r, column=2)
    target_label.font = Font(bold=True)
    target_value.font = Font(bold=True)
    target_value.number_format = currency_format
    r += 1
    ws.cell(row=r, column=1, value="Alignment with target balance?")
    ws.cell(row=r, column=2, value=f"=B{target_balance_row}=B{approved_projected_balance_row}")

    r += 2

    ws.cell(row=r, column=1, value="FreeAgent's current balance for Account B (Clean), as of this fetch:")
    ws.cell(row=r, column=2, value=(f"=SUM('{clean_sheet_name}'!${amount_col}:${amount_col})"))
    ws.cell(row=r, column=2).number_format = currency_format

    r += 2

    ws.cell(row=r, column=1, value="Number of transactions").font = Font(bold=True)
    r += 1

    ws.cell(row=r, column=1, value="In Account A (Polluted):")
    ws.cell(row=r, column=2, value=f"=COUNTA('{polluted_sheet_name}'!$A:$A)-1")
    ws.cell(row=r, column=2).number_format = count_format
    total_transactions_row = r
    r += 1

    ws.cell(row=r, column=1, value="Approved for deletion:")
    ws.cell(row=r, column=2, value=(
        f"=COUNTIFS('{polluted_sheet_name}'!$O:$O,\"Y\","
        f"'{polluted_sheet_name}'!$N:$N,\"matched\")"
    ))
    ws.cell(row=r, column=2).number_format = count_format
    approved_transactions_row = r
    r += 1

    ws.cell(row=r, column=1, value="Remaining after approved deletions:")
    ws.cell(row=r, column=2, value=f"=B{total_transactions_row}-B{approved_transactions_row}")
    ws.cell(row=r, column=2).number_format = count_format
    r += 1

    return r


def add_instructions_sheet(wb, polluted_earliest, clean_earliest, polluted_current_balance, polluted_sheet_name, clean_sheet_name):
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 100
    lines = [
        ("How to use this workbook", True),
        ("", False),
        ("1. 'Account A (Polluted)' and 'Account B (Clean)' are raw pulls from FreeAgent — one row per transaction, unedited.", False),
        ("2. On Account A, 'match_count' counts how many transactions in Account B share the same date + amount (COUNTIFS).", False),
        ("   'match_status' turns that into: unmatched / matched / AMBIGUOUS - review.", False),
        ("3. 'matched' (exactly one same-day, same-amount transaction in Account B) is the confident case: almost", False),
        ("   certainly a duplicate injected by the misconfigured feed.", False),
        ("4. 'AMBIGUOUS - review' means more than one candidate on that date/amount in Account B (or vice versa) —", False),
        ("   filter Account B by that date to see the candidates yourself before deciding.", False),
        ("5. Check 'explanation_type', 'is_deletable' and 'is_locked' before approving anything that's explained:", False),
        ("   - is_deletable = N: FreeAgent won't let the explanation be removed as-is.", False),
        ("   - explanation_type contains 'Invoice Receipt': deleting can re-open an invoice as unpaid and email the customer.", False),
        ("   - is_locked = Y (often VAT-related): removing the explanation will defer the change to your next open VAT period.", False),
        ("6. Set 'approve_delete' to Y on the rows (on Account A only) you want deleted, then save this file.", False),
        ("7. Run: python delete_transactions.py <this file>          (dry run — logs what it WOULD do)", False),
        ("   Then: python delete_transactions.py <this file> --live  (only after checking the dry-run log)", False),
        ("", False),
        (f"Account A (Polluted) earliest transaction date: {polluted_earliest}", False),
        (f"Account B (Clean) earliest transaction date:    {clean_earliest}", False),
        ("If B's earliest date doesn't reach back far enough to cover the whole period the wrong feed was active,", False),
        ("anything in A older than that falls outside what this diff can safely catch.", False),
    ]
    for i, (text, is_title) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if is_title:
            cell.font = Font(bold=True, size=13)

    add_balance_reconciliation(ws, len(lines) + 2, polluted_current_balance, polluted_sheet_name, clean_sheet_name)


def main():
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    polluted_id = os.environ["ACCOUNT_POLLUTED_ID"]
    clean_id = os.environ["ACCOUNT_CLEAN_ID"]

    from_cache = "--from-cache" in sys.argv

    polluted_rows = fetch_and_cache(polluted_id, "polluted", from_cache=from_cache)
    clean_rows = fetch_and_cache(clean_id, "clean", from_cache=from_cache)
    polluted_current_balance = get_current_balance(polluted_id, from_cache=from_cache)

    wb = openpyxl.Workbook()
    ws_polluted = wb.active
    ws_polluted.title = "Account A (Polluted)"
    ws_clean = wb.create_sheet("Account B (Clean)")

    write_raw_sheet(ws_polluted, polluted_rows)
    write_raw_sheet(ws_clean, clean_rows)
    add_match_and_approval_columns(ws_polluted, len(polluted_rows), "Account B (Clean)")

    polluted_earliest = min((r["dated_on"] for r in polluted_rows), default="n/a")
    clean_earliest = min((r["dated_on"] for r in clean_rows), default="n/a")
    add_instructions_sheet(
        wb, polluted_earliest, clean_earliest,
        polluted_current_balance, ws_polluted.title,
        ws_clean.title,
    )

    out_path = f"review_{ts}.xlsx"
    wb.save(out_path)

    print(f"\nWrote {out_path}")
    print(f"Account A (Polluted) earliest date: {polluted_earliest}")
    print(f"Account B (Clean) earliest date:    {clean_earliest}")
    print(
        "\nOpen the file in Excel — it needs one open/recalculate pass before "
        "match_count/match_status show values instead of 0. Read the "
        "'Instructions' tab first."
    )


if __name__ == "__main__":
    main()
