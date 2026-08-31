"""
delete_transactions.py — executes deletions for rows approved in the review
workbook.

Usage:
    python delete_transactions.py review_<timestamp>.xlsx            # dry run
    python delete_transactions.py review_<timestamp>.xlsx --live     # for real

Only acts on rows in "Account A (Polluted)" where:
    - match_status == "matched"     (not "unmatched" or "AMBIGUOUS - review")
    - approve_delete is "Y"          (case-insensitive)

For each row it re-fetches the LIVE transaction/explanation state right
before acting (the workbook may be stale by the time you run this), skips
anything not currently deletable or flagged as an Invoice Receipt, removes
the explanation if present, then deletes the transaction. Every action is
logged to delete_log_<timestamp>.csv so a partial run can be picked back up
and audited.
"""
import csv
import sys
from datetime import datetime, timezone

import openpyxl

import freeagent_client as fa


def load_approved_rows(xlsx_path):
    # data_only=True reads cached formula results. If you've never opened
    # this file in Excel/LibreOffice since editing approve_delete, those
    # cells may not be recalculated yet — open and save it once first.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Account A (Polluted)"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["url"]] is None:
            continue
        approve = str(row[idx.get("approve_delete", -1)] or "").strip().upper()
        status = str(row[idx.get("match_status", -1)] or "").strip().lower()
        if approve == "Y" and status == "matched":
            rows.append({
                "url": row[idx["url"]],
                "dated_on": row[idx["dated_on"]],
                "amount": row[idx["amount"]],
                "description": row[idx["description"]],
            })
    return rows


def process(rows, live, log_path):
    with open(log_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["url", "dated_on", "amount", "description", "action", "result", "detail"])

        for row in rows:
            url, dated_on, amount, desc = row["url"], row["dated_on"], row["amount"], row["description"]
            print(f"\n{url}  {dated_on}  {amount}  {desc}")

            try:
                live_txn = fa.get_transaction(url)
            except Exception as e:
                writer.writerow([url, dated_on, amount, desc, "fetch", "ERROR", str(e)])
                print(f"  Could not fetch current state, skipping: {e}")
                continue

            explanations = live_txn.get("bank_transaction_explanations") or []

            if explanations:
                exp_url = explanations[0]["url"]
                detail = fa.get_explanation(exp_url)
                exp_type = detail.get("type") or detail.get("entry_type") or ""

                if not detail.get("is_deletable", True):
                    reason = detail.get("locked_reason") or "not deletable"
                    print(f"  SKIP: explanation not deletable ({reason})")
                    writer.writerow([url, dated_on, amount, desc, "remove_explanation", "SKIPPED", reason])
                    continue

                if "invoice receipt" in exp_type.lower():
                    print(
                        f"  SKIP: explanation type is '{exp_type}' — deleting could re-open the "
                        f"linked invoice as unpaid and trigger reminder emails. Handle manually."
                    )
                    writer.writerow([url, dated_on, amount, desc, "remove_explanation", "SKIPPED", f"type={exp_type}"])
                    continue

                if live:
                    try:
                        fa.delete_explanation(exp_url)
                        writer.writerow([url, dated_on, amount, desc, "remove_explanation", "OK", exp_type])
                        print(f"  Removed explanation ({exp_type})")
                    except Exception as e:
                        writer.writerow([url, dated_on, amount, desc, "remove_explanation", "ERROR", str(e)])
                        print(f"  ERROR removing explanation: {e}")
                        continue
                else:
                    print(f"  [dry run] would remove explanation ({exp_type})")
                    writer.writerow([url, dated_on, amount, desc, "remove_explanation", "DRY_RUN", exp_type])

            if live:
                try:
                    fa.delete_transaction(url)
                    writer.writerow([url, dated_on, amount, desc, "delete_transaction", "OK", ""])
                    print("  Deleted transaction")
                except Exception as e:
                    writer.writerow([url, dated_on, amount, desc, "delete_transaction", "ERROR", str(e)])
                    print(f"  ERROR deleting transaction: {e}")
            else:
                print("  [dry run] would delete transaction")
                writer.writerow([url, dated_on, amount, desc, "delete_transaction", "DRY_RUN", ""])


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    live = "--live" in sys.argv

    rows = load_approved_rows(xlsx_path)
    print(f"{len(rows)} approved + matched rows found in {xlsx_path}")
    if not rows:
        return

    if live:
        confirm = input(
            f"\nThis will PERMANENTLY delete {len(rows)} transactions from your "
            f"live FreeAgent account. Type DELETE to continue: "
        )
        if confirm != "DELETE":
            print("Aborted.")
            return
    else:
        print("DRY RUN — nothing will be changed. Re-run with --live to actually delete.\n")

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    log_path = f"delete_log_{ts}.csv"
    process(rows, live, log_path)
    print(f"\nLog written to {log_path}")


if __name__ == "__main__":
    main()
